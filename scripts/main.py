#!/usr/bin/env python
# _*_ encoding: utf-8 _*_
# ---------------------------------------
# Created by: Jlfme<jlfgeek@gmail.com>
# Created on: 2017-10-21 18:59:18
# ---------------------------------------


import os

from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter
from openpyxl.styles.fonts import Font


def int_to_letter(column):
        """ 将数字转化为excel的字母列，　例如　1 > A, 30 > AD
        # More information please visit http://support.microsoft.com/kb/833402
        """
        result = ""
        alpha = column // 27
        remainder = column - (alpha * 26)
        if alpha > 0:
            result = chr(alpha + 64)
        if remainder > 0:
            result += chr(remainder + 64)
        return result


def write_data(sheet, row, data_list):
    font = Font(name='Times New Roman', size=10)

    coordinates = 'K{}:CF{}'.format(row, row)
    cells = sheet[coordinates][0]
    for i, cell in enumerate(cells):
        cell.value = data_list[i]
        cell.font = font


def format_sheet(sheet):
    """根据内容长度自动设置宽度
    """
    column_width_dict = {}
    for i in range(11, sheet.max_column + 1):
        column = get_column_letter(i)
        coordinates = '{column}3:{column}{max_row}'.format(column=column, max_row=sheet.max_row)
        withs = [len(cell[0].value) for cell in sheet[coordinates] if cell[0].value is not None]

        # 取最大长度的值
        column_width_dict[column] = float(max(withs) + 1)

    # 设置单元格的宽度
    for k, v in column_width_dict.items():
        sheet.column_dimensions[k].width = v


def read_excel(file_path):
    print('正在读取文件:', file_path)
    print('>>>正在处理...')

    try:
        wb = load_workbook(file_path, read_only=False, data_only=True)
        sheet = wb.active

        # generate coordinates
        total = 0
        coordinates = 'J2:J{}'.format(sheet.max_row)
        cells = sheet[coordinates]
        for cell in cells:
            cell = cell[0]
            data = cell.value
            if data is not None:
                data_list = cell.value.split(';')
                if len(data_list) > 10:
                    total += 1
                    write_data(sheet, cell.row, data_list)

        # 设置单元格宽度
        format_sheet(sheet)
    except Exception as e:
        print('警告: 处理文件出错了!', file_path, e)
        os.system('pause')
    else:
        print('>>>处理完毕, 处理总行数:', total)
        print('正在保存文件', file_path)

        # save excel
        out_dir = os.path.join(os.getcwd(), 'out')
        if not os.path.exists(out_dir):
            os.mkdir(out_dir)

        new_file_path = os.path.join(out_dir, os.path.basename(file_path))
        wb.save(new_file_path)
        print("------------------------------------------------------------------")


def find_excel_files():
    base_dir = os.path.join(os.getcwd(), 'files')
    if not os.path.exists(base_dir):
        print('警告:  <files>文件夹不存才, 正在创建...')
        os.mkdir(base_dir)
        print('提示:  请将所有需要处理的excel文件放到<files>目录中, 然后重新运行程序!')
        os.system('pause')
    else:
        print('正在读取目录<files>中的所有excel文件...')
        files = os.listdir(base_dir)
        print('>>>待处理文件分析完成, 总数:', len(files))

        print("------------------------------------------------------------------")

        file_path_list = [os.path.join(base_dir, filename) for filename in files]
        for file_path in file_path_list:
            if '.xls' in file_path or '.xlsx' in file_path:
                read_excel(file_path)

    print('\n<--------------------------所有文件操作完成------------------------->')
    os.system('pause')


if __name__ == '__main__':
    find_excel_files()
